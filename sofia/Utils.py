import math
import os
from openpyxl import Workbook
import string
from bs4 import BeautifulSoup
import enchant

class Quantifiers:
    def __init__(self, quantifiers, entities, sentence):
        self.sentence= sentence
        self.quantifiers= quantifiers
        self.entities= entities
        self.matchDict= self.match()

    def getIndex(self, term):
        termTokens= term.split(' ')
        tokens= self.sentence. split(' ')
        exception= True
        i=0
        while exception:
            try:
                start= tokens.index(termTokens[i])-i
                exception= False
            except:
                exception = True
                i+=1
        end= tokens.index(termTokens[-1])
        exception = True
        i = 0
        while exception:
            try:
                end = tokens.index(termTokens[-i])+i
                exception= False
            except:
                exception = True
                i-=1
        return start, end

    def getDistance(self, s1, t1, s2, t2):
        diff = -float('inf')
        if s2<= t1:
            if t2>=s1:
                return 0
            return s1-t2
        return s2-t1

    def chooseClosest(self, term1, term2, target):
        start, end= self.getIndex(target)
        s1, t1= self.getIndex(term1)
        s2, t2= self.getIndex(term2)
        diff1= self.getDistance(s1, t1, start, end)
        diff2 =self.getDistance(s2, t2, start, end)
        if diff1<diff2:
            return term1
        return term2

    def getClosest(self, terms, target):
        if len(terms)==1:
            return terms[0]
        bestTerm= terms[0]
        for term in terms[1:]:
            bestTerm= self.chooseClosest(bestTerm, term, target)
        return bestTerm

    def match(self):
        if len(self.entities)==0:
            return {}
        Dic={}
        for quant in self.quantifiers:
            best= self.getClosest(self.entities, quant)
            if best in Dic:
                Dic[best]+= [quant]
            else:
                Dic[best]= [quant]
        matchDic={}
        for entity in list(Dic.keys()):
            quants= Dic[entity]
            if len(quants)==1:
                matchDic[entity]= quants[0]
            else:
                matchDic[entity]= self.getClosest(quants, entity)
        return matchDic

    def getQuantifier(self, entity):
        if entity not in self.matchDict:
            return ''
        return  self.matchDict[entity]

class FileReader:
    import xml.etree.ElementTree as ET

    def __init__(self, dataDir='None'):
        self.dir = os.path.dirname(os.getcwd())+'/data/'
        self.d=  enchant.Dict("en_US")
        if dataDir != 'None':
            self.dir= dataDir+ '/dataFiles/'

    def readFile(self, file):
        f= open(self.dir+file)
        text= f.read()
        f.close()
        fText= self.cleanText(text)
        f = open(self.dir + 'output/' + file, 'w')
        f.write(fText)
        f.close()

    def cleanText(self, text):
        sentences = text.split('\n')
        newText = ""
        for s in sentences:
            s = s.strip('\n')
            s = s.strip(' ')
            newSent = self.textToAscii(s)
            if len(newSent.split(' ')) > 25:
                if '\n' in newSent:
                    i = newSent.index('\n')
                    newText += newSent[:i] + '. ' + newSent[i:]
            elif len(newSent.split(' ')) > 4:
                newText += newSent + '\n'
        return newText

    def textToAscii(self, input):
        newText = ''
        for letter in input:
            if ord(letter) < 128:
                if letter == '\n':
                    newText += ' '
                else:
                    newText += letter
        words = newText.split(' ')
        newWords = []
        for w in words:
            if len(w) > 10:
                if self.d.check(w):
                    newWords.append(w)
            else:
                newWords.append(w)
        return string.join(newWords, ' ')

    def removeAnnotations(self, text, limitLeft, limitRight):
        annotationSpans={}
        newText= ''
        i=0
        new_i=0
        while i< len(text):
            currLetter= text[i]
            if limitLeft== currLetter:
                currAnnotation=currLetter
                loop= True
                loop2 = True
                start=0
                end=1
                while loop:
                    i += 1
                    #new_i+=1
                    currLetter = text[i]
                    currAnnotation+= currLetter
                    if limitRight== currLetter and loop2:
                        i+=1
                        #new_i += 1
                        currLetter = text[i]
                        currAnnotation += currLetter
                        start = new_i
                        while loop2:
                            if limitLeft== currLetter:
                                loop2=False
                                end= new_i
                                new_i-=1
                            else:
                                newText += currLetter
                                i += 1
                                new_i += 1
                                currLetter = text[i]
                                currAnnotation += currLetter
                    elif limitRight== currLetter:
                        #i += 1
                        # currLetter = text[i]
                        # currAnnotation += currLetter
                        loop= False
                        annotationSpans[(start, end)] = currAnnotation
            else:
                newText+= currLetter
            i += 1
            new_i += 1
        return newText, annotationSpans

    def run(self, files):
        for file in files:
            print(file)
            f=open(file)
            text_i= f.read()
            f.close()
            index = text_i.index('<MAKEINSTANCE')
            text_i=text_i[:index]
            newT, spans = self.removeAnnotations(' ', text_i, '<', '>')
            f = open('/Users/evangeliaspiliopoulou/Desktop/Structured_Learner/JointModel/data/TimebankDense/Annotations/' + file[:-4], 'w')
            k = list(spans.keys())
            k.sort()
            for key in k:
                item = spans[key]
                if len(item) > 1:
                    if '\n' in item:
                        i = item.index('\n')
                        item = item[:i] + item[i + 1:]
                    f.write(str(item) + '\t' + str(key) + '\n')
            f.close()


    def readXmlFile(self, file):
        f = open(self.dir + file)
        data = f.read()
        f.close()
        soup = BeautifulSoup(data)
        text = soup.get_text()
        fText = self.cleanText(text)
        fileName= file.split('.')[0]
        f = open(self.dir + 'output/' + fileName+'.txt', 'w')
        f.write(fText)
        f.close()


    def sortDict(self, dict):
        dictList=[]
        for key, value in sorted(iter(dict.items()), key=lambda k_v: (k_v[1], k_v[0]), reverse=True):
            dictList.append((key, value))
        return dictList

    def readFiles(self, format):
        files= os.listdir(self.dir)
        for file in files:
            if file!= '.DS_Store' and file!= 'output':
                print("processing file " + str(file))
                if format== 'txt':
                    self.readFile(file)

                else:
                    try:
                        self.readXmlFile(file)
                    except:
                        print("EXCEPTION!")
                        self.readFile(file)

    def findTerm(self, myList, phrase, key):
        index=0
        while index < len(myList):
            phraseIndex=0
            while myList[index][key]== phrase[phraseIndex]:
                index+=1
                phraseIndex+=1
                if phraseIndex== len(phrase):
                    return myList[index-phraseIndex: index]
            if phraseIndex<len(phrase):
                index-=phraseIndex-1
        return []



class ExcelWriter:
    def __init__(self, list_sheets):
        self.list_sheets= list_sheets
        self.wb= Workbook()
        ws = self.wb.active
        self.sheets=[ws]
        self.currIndex=[1]
        ss_sheet = self.wb.get_sheet_by_name('Sheet')
        ss_sheet.title = list_sheets[0]
        s= 1
        while len(list_sheets)-s>0:
            name= list_sheets[s]
            newWs = self.wb.create_sheet(name)
            self.sheets.append(newWs)
            self.currIndex.append(1)
            s+=1

    def writeRow(self, sheet, rowElements):
        letters = list(string.ascii_uppercase)
        sheetIndex= self.list_sheets.index(sheet)
        ws= self.sheets[sheetIndex]
        currIndex= self.currIndex[sheetIndex]
        for h in range(len(rowElements)):
            header= rowElements[h]
            letter= letters[h]
            ws[letter+str(currIndex)]= header
        self.currIndex[sheetIndex]+=1

    def readRoW(self, sheet, index):
        letters = list(string.ascii_uppercase)
        sheetIndex = self.list_sheets.index(sheet)
        ws= self.sheets[sheetIndex]
        row={}
        curr=0
        letter= 'A'
        value= ws[letter+'1'].value
        while value!= '':
            cell= ws[letter+ str(index)]
            row[value]= cell
            curr+=1
            letter= letters[curr]
            value = ws[letter + '1'].value
        return row

    def getIndex(self, sheet):
        sheetIndex = self.list_sheets.index(sheet)
        index= self.currIndex[sheetIndex]
        return index


    def saveExcelFile(self, dir, name):
        self.wb.save(dir+'/'+ name)

    def getIndexFromSpan(self, localIndex, spanList):
        if spanList==0:
            return ""
        index=""
        for span in spanList:
            i= localIndex[span]
            if i!= "":
                index+= i+ ', '
        index= index.strip(', ')
        return index

def getIndexFromSpan(localIndex, spanList):
    if spanList==0:
        return ""
    index=""
    for span in spanList:
        i= localIndex[span]
        if i!= "":
            index+= i+ ', '
    index= index.strip(', ')
    return index